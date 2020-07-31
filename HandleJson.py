import json
import csv
import configparser
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from time import strftime
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl.styles import Font

wb = Workbook()
del wb['Sheet']
json_root_path = "jsons"

def get_keys(dl, keys_list):
    if isinstance(dl, dict):
        keys_list += dl.keys()
        map(lambda x: get_keys(x, keys_list), dl.values())
    elif isinstance(dl, list):
        map(lambda x: get_keys(x, keys_list), dl)


def HandleParseFiles():
    config = configparser.ConfigParser()
    config.read('HandleJson.ini')
    count = int(config['counts']['fileParse'])
    with open(json_root_path+'/ParseFiles.json') as json_file:
        data = json.load(json_file)
    sheet_name = data['Type']
    ws = wb.create_sheet(sheet_name)  # insert at the end (default)
    ws.title = data['Type']
    count = min(len(data['File Names']),count)
    for i in range(0, count):
        ws['A' + str(i + 3)] = data['File Names'][i]
        ws['B' + str(i + 3)] = str(data['Time Spend'][i]) + " ms"
        ws['B' + str(i + 3)].alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 220
    ws.column_dimensions['B'].width = 20

    ws.merge_cells('A1:B1')
    ws['A1'] = "Top " + str(count) + " Files Taking Longgest Time To Parse (Compiler Frontend)"
    ws['A2'] = "File Names"
    ws['B2'] = "Time Spend"

    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['B2'].alignment = Alignment(horizontal='center')
    ws['A1'].font = Font(bold=True, size=15)
    ws['A2'].font = Font(bold=True)
    ws['B2'].font = Font(bold=True)


def HandleCodegenFiles():
    config = configparser.ConfigParser()
    config.read('HandleJson.ini')
    count = int(config['counts']['fileCodegen'])
    with open(json_root_path+'/CodegenFiles.json') as json_file:
        data = json.load(json_file)
    sheet_name = data['Type']
    ws = wb.create_sheet(sheet_name)  # insert at the end (default)
    ws.title = data['Type']
    count = min(len(data['File Names']),count)
    for i in range(0, count):
        ws['A' + str(i + 3)] = data['File Names'][i]
        ws['B' + str(i + 3)] = str(data['Time Spend'][i])+" ms"
        ws['B' + str(i + 3)].alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 220
    ws.column_dimensions['B'].width = 20

    ws.merge_cells('A1:B1')
    ws['A1'] = "Top " + str(count) + " Files Taking Longgest Time To Codegen (Compiler Backend)"
    ws['A2'] = "File Names"
    ws['B2'] = "Time Spend"

    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['B2'].alignment = Alignment(horizontal='center')
    ws['A1'].font = Font(bold=True, size=15)
    ws['A2'].font = Font(bold=True)
    ws['B2'].font = Font(bold=True)

def HandleWhloeCompilationInfo():
    with open(json_root_path+'/WholeCompilation.json') as json_file:
        data = json.load(json_file)
    sheet_name = data['Type']
    ws = wb.create_sheet(sheet_name)  # insert at the end (default)
    ws.merge_cells('A1:B1')
    ws.title = data['Type']
    ws['A1'] = "Whole Compilation Information"
    ws['A2'] = "Total Compilation Time"
    ws['B2'] = "{:.2f}".format(data['Total Time'] / 1000000.0) + " s"
    ws['A3'] = "Parsing Time"
    ws['B3'] =  "{:.2f}".format(data['Parsing Time'] / 1000000.0) + " s"
    ws['A4'] = "Code Generation & Optimization Time"
    ws['B4'] = "{:.2f}".format(data['Code Gen&Opts Time'] / 1000000.0) + " s"

    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 20
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A3'].alignment = Alignment(horizontal='center')
    ws['A4'].alignment = Alignment(horizontal='center')

    ws['A1'].font = Font(bold=True, size=15)
    ws['A2'].font = Font(bold=True)
    ws['A3'].font = Font(bold=True)
    ws['A4'].font = Font(bold=True)

def HandleTemplatesInstantiations():
    config = configparser.ConfigParser()
    config.read('HandleJson.ini')
    count = int(config['counts']['template'])
    with open(json_root_path + '/Instantiations.json') as json_file:
        data = json.load(json_file)
    keys = []
    get_keys(data, keys)
    sheet_name = data['Type']
    ws = wb.create_sheet(sheet_name)  # insert at the end (default)
    ws.title = data['Type']
    count = min(len(data['File Names']),count)
    for i in range(0, count):
        ws['A' + str(i + 3)] = data['File Names'][i]
        ws['B' + str(i + 3)] = str(data['Time Spend'][i]) + " ms"
        ws['B' + str(i + 3)].alignment = Alignment(horizontal='center')
        ws['C' + str(i + 3)] = str(data['Instantiation times'][i])
        ws['C' + str(i + 3)].alignment = Alignment(horizontal='center')
        ws['D' + str(i + 3)] = str(data['Average time'][i]) + " ms"
        ws['D' + str(i + 3)].alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 220
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    ws.merge_cells('A1:D1')
    ws['A1'] = "Top " + str(count) + " Templates Taking Longgest Time To Instantiate"
    ws['A2'] = "File Names"
    ws['B2'] = "Time Spend"
    ws['C2'] = "Instantiation times"
    ws['D2'] = "Average time"

    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['B2'].alignment = Alignment(horizontal='center')
    ws['C2'].alignment = Alignment(horizontal='center')
    ws['D2'].alignment = Alignment(horizontal='center')

    ws['A1'].font = Font(bold=True, size=15)
    ws['A2'].font = Font(bold=True)
    ws['B2'].font = Font(bold=True)
    ws['C2'].font = Font(bold=True)
    ws['D2'].font = Font(bold=True)

def HandleTemplatesInstantiationsSets():
    config = configparser.ConfigParser()
    config.read('HandleJson.ini')
    count = int(config['counts']['template'])
    with open(json_root_path + '/InstantiationsSets.json') as json_file:
        data = json.load(json_file)
    keys = []
    get_keys(data, keys)
    sheet_name = data['Type']
    ws = wb.create_sheet(sheet_name)  # insert at the end (default)
    ws.title = data['Type']
    count = min(len(data['File Names']),count)
    for i in range(0, count):
        ws['A' + str(i + 3)] = data['File Names'][i]
        ws['B' + str(i + 3)] = str(data['Time Spend'][i]) + " ms"
        ws['B' + str(i + 3)].alignment = Alignment(horizontal='center')
        ws['C' + str(i + 3)] = str(data['Instantiation times'][i])
        ws['C' + str(i + 3)].alignment = Alignment(horizontal='center')
        ws['D' + str(i + 3)] = str(data['Average time'][i]) + " ms"
        ws['D' + str(i + 3)].alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    ws.merge_cells('A1:D1')
    ws['A1'] = "Top " + str(count) + " Template Sets Taking Longgest Time To Instantiate"
    ws['A2'] = "File Names"
    ws['B2'] = "Time Spend"
    ws['C2'] = "Instantiation times"
    ws['D2'] = "Average time"

    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['B2'].alignment = Alignment(horizontal='center')
    ws['C2'].alignment = Alignment(horizontal='center')
    ws['D2'].alignment = Alignment(horizontal='center')

    ws['A1'].font = Font(bold=True, size=15)
    ws['A2'].font = Font(bold=True)
    ws['B2'].font = Font(bold=True)
    ws['C2'].font = Font(bold=True)
    ws['D2'].font = Font(bold=True)

def HandleFunctions():
    config = configparser.ConfigParser()
    config.read('HandleJson.ini')
    count = int(config['counts']['function'])
    with open(json_root_path + '/Functions.json') as json_file:
        data = json.load(json_file)
    keys = []
    get_keys(data, keys)
    sheet_name = data['Type']
    ws = wb.create_sheet(sheet_name)  # insert at the end (default)
    ws.title = data['Type']
    count = min(len(data['Function Names']), count)
    for i in range(0, count):
        ws['A' + str(i + 3)] = data['Function Names'][i]
        ws['B' + str(i + 3)] = str(data['Time Spend'][i]) + " ms"
        ws['B' + str(i + 3)].alignment = Alignment(horizontal='center')
        ws['C' + str(i + 3)] = str(data['File Names'][i])
        ws['C' + str(i + 3)].alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 220

    ws.merge_cells('A1:C1')
    ws['A1'] = "Top " + str(count) + " Functions Taking Longgest Time To Compile"
    ws['A2'] = "File Names"
    ws['B2'] = "Time Spend"
    ws['C2'] = "File Names"

    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['B2'].alignment = Alignment(horizontal='center')
    ws['C2'].alignment = Alignment(horizontal='center')

    ws['A1'].font = Font(bold=True, size=15)
    ws['A2'].font = Font(bold=True)
    ws['B2'].font = Font(bold=True)
    ws['C2'].font = Font(bold=True)

def HandleFunctionSets():
    config = configparser.ConfigParser()
    config.read('HandleJson.ini')
    count = int(config['counts']['function'])
    with open(json_root_path + '/FunctionSets.json') as json_file:
        data = json.load(json_file)
    keys = []
    get_keys(data, keys)
    sheet_name = data['Type']
    ws = wb.create_sheet(sheet_name)  # insert at the end (default)
    ws.title = data['Type']
    count = min(len(data['File Names']), count)
    for i in range(0, count):
        ws['A' + str(i + 3)] = data['File Names'][i]
        ws['B' + str(i + 3)] = str(data['Time Spend'][i]) + " ms"
        ws['B' + str(i + 3)].alignment = Alignment(horizontal='center')
        ws['C' + str(i + 3)] = str(data['Compile/optimization times'][i])
        ws['C' + str(i + 3)].alignment = Alignment(horizontal='center')
        ws['D' + str(i + 3)] = str(data['Average time'][i]) + " ms"
        ws['D' + str(i + 3)].alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    ws.merge_cells('A1:D1')
    ws['A1'] = "Top " + str(count) + " Function Sets Taking Longgest Time To Compile"
    ws['A2'] = "File Names"
    ws['B2'] = "Time Spend"
    ws['C2'] = "Instantiation times"
    ws['D2'] = "Average time"

    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['B2'].alignment = Alignment(horizontal='center')
    ws['C2'].alignment = Alignment(horizontal='center')
    ws['D2'].alignment = Alignment(horizontal='center')

    ws['A1'].font = Font(bold=True, size=15)
    ws['A2'].font = Font(bold=True)
    ws['B2'].font = Font(bold=True)
    ws['C2'].font = Font(bold=True)
    ws['D2'].font = Font(bold=True)

def HandleExpensiveHeaders():
    config = configparser.ConfigParser()
    config.read('HandleJson.ini')
    count = int(config['counts']['header'])
    with open(json_root_path + '/ExpensiveHeaders.json') as json_file:
        data = json.load(json_file)
    keys = []
    get_keys(data, keys)
    sheet_name = data['Type']
    ws = wb.create_sheet(sheet_name)  # insert at the end (default)
    ws.title = data['Type']
    count = min(len(data['Header Names']), count)
    for i in range(0, count):
        ws['A' + str(i + 3)] = data['Header Names'][i]
        ws['B' + str(i + 3)] = str(data['Total Time'][i]) + " ms"
        ws['B' + str(i + 3)].alignment = Alignment(horizontal='center')
        ws['C' + str(i + 3)] = str(data['Included Times'][i])
        ws['C' + str(i + 3)].alignment = Alignment(horizontal='center')
        ws['D' + str(i + 3)] = str(data['Average Time'][i]) + " ms"
        ws['D' + str(i + 3)].alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 120
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    ws.merge_cells('A1:D1')
    ws['A1'] = "Top " + str(count) + " Expensive Headers"
    ws['A2'] = "Header Names"
    ws['B2'] = "Total Time"
    ws['C2'] = "Included Times"
    ws['D2'] = "Average time"

    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['B2'].alignment = Alignment(horizontal='center')
    ws['C2'].alignment = Alignment(horizontal='center')
    ws['D2'].alignment = Alignment(horizontal='center')

    ws['A1'].font = Font(bold=True, size=15)
    ws['A2'].font = Font(bold=True)
    ws['B2'].font = Font(bold=True)
    ws['C2'].font = Font(bold=True)
    ws['D2'].font = Font(bold=True)

if __name__ == '__main__':
    HandleWhloeCompilationInfo()
    HandleParseFiles()
    HandleCodegenFiles()
    HandleTemplatesInstantiations()
    HandleTemplatesInstantiationsSets()
    HandleFunctions()
    HandleFunctionSets()
    HandleExpensiveHeaders()
    time = datetime.now().strftime('%Y-%m-%d %H点%M分%S秒')
    wb.save('ClangAnalyzeResult ' + time + '.xlsx')
